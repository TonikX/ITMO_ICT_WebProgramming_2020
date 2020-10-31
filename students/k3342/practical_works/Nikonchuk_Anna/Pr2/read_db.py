from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('/home/nikon-cook/Documents/МИТМО/веб-прога/minos_db.xlsx')

print(wb.get_sheet_names())
Publishers_data = wb.get_sheet_by_name('Publishers')
#print(Publishers_data.dimensions, Publishers_data.max_row, Publishers_data.max_column)
Authors_data = wb.get_sheet_by_name('Authors')
Books_data = wb.get_sheet_by_name('Books')
Directors_data = wb.get_sheet_by_name('Directors')


'''
p_df = pd.DataFrame()
p_rows = Publishers_data.iter_rows()
p_first_row = next(p_rows)
p_sec_row = next(p_rows)
p_headings = [c.value for c in p_first_row]
p_row1 = [c.value for c in p_sec_row]
p_headings
print(p_headings, p_row1)
'''
'''
df = []
for i in range(Publishers_data.max_row):
	df.append([c.value for c in next(p_rows)])

publisher_df = pd.DataFrame(df[1:])
publisher_df.columns = df[0]
print(publisher_df)
'''
'''
a_rows = Authors_data.iter_rows()
a_first_row = next(a_rows)
a_headings = [c.value for c in a_first_row]
a_headings
'''

def make_dataframe(data_sheet_name):
	df = []
	rows = data_sheet_name.iter_rows()
	for i in range(data_sheet_name.max_row):
		df.append([c.value for c in next(rows)])

	res_df = pd.DataFrame(df[1:])
	res_df.columns = df[0]
	return res_df

publishers_df = make_dataframe(Publishers_data)
authors_df = make_dataframe(Authors_data)
books_df = make_dataframe(Books_data)
directors_df = make_dataframe(Directors_data)

print(publishers_df.shape[0])
print(publishers_df.index, '\n')
#print(publishers_df.iloc[0])


def create_comands_txt(str_model_class, df, str_file_name, foreign):
	f = open(str_file_name, 'w')
	counter = 0
	#comands = ''
	headings = list(df.columns)
	for i in range(df.shape[0]):
		one_command_str = str_model_class + '.objects.create('
		#print(df.iloc[i][headings[0]])
		for j in range(df.shape[1]):
			#print('   ',headings[j],' - ', str(df.iloc[i][headings[j]]).lstrip().rstrip())
			if df.iloc[i][headings[j]] != None and df.iloc[i][headings[j]] != "None":
				if headings[j] not in foreign:
					one_command_str = one_command_str + headings[j] + '=\"' + str(df.iloc[i][headings[j]]).lstrip().rstrip() + '\",'
				else:
					#print('else')
					comma_pos = str(df.iloc[i][headings[j]]).find(',') 
					if comma_pos > 0 and headings[j][-1]=='s':
						str2 = headings[j][:-1].capitalize() + '.objects.get(sort_name=\"'
						str_with_comma = str(df.iloc[i][headings[j]]).lstrip().rstrip()
						one_command_str = one_command_str + headings[j] + '=[' + str2 +\
						str_with_comma[0:comma_pos] + ')\",' + str2 + str_with_comma[comma_pos:] + '\")],'
					else:
						one_command_str = one_command_str + headings[j] + '=' +\
						headings[j].capitalize() + '.objects.get(name=\"' + \
						str(df.iloc[i][headings[j]]).lstrip().rstrip() + '\"),'
		one_command_str = one_command_str[:-1] + ')\n'
		#comands += one_command_str
		f.write(one_command_str)
		counter += 1
	f.close()
	print('generated '+str(counter)+' commands')
	return 0

#create_comands_txt('Publisher', publishers_df, 'publisher_comands.txt')
#create_comands_txt('Author', authors_df, 'author_comands.txt')
#create_comands_txt('Book', books_df, 'book_comands.txt')
#create_comands_txt('Director', directors_df, 'director_comands.txt')

count = books_df.shape[0]
n = pd.Series([str(i) for i in range(190)])
type_cover = list(books_df.type_cover)
name = list(books_df.orig_lang_name)

res = []
for i in range(count):
	str1 = ''
	if type_cover[i]!=None:
		str1 += type_cover[i] + '_'
	str1 += n[i] + '_'
	str1 += name[i].replace(' ','_')
	#print(str1)
	res.append(str1)
books_df['b_slug'] = pd.Series(res)
# print(books_df.b_slug)
create_comands_txt('Book', books_df, 'book_comands.txt', ["authors", "publisher"])

#print(n + c + books_df.type_cover + c)
#print(pd.Series(books_df.index))

'''
books_df['b_slug'] = books_df.indexes + '_' + books_df.orig_lang_name

for i in range(books_df.shape[0]):
	str1 = str(i) + '_' + books_df.iloc(i)[0].orig_lang_name.replace(' ','_') + \
		'_' + books_df.iloc(i)[0].type_cover + '_' + \
		str(books_df.iloc(i)[0].publication_date)[:4]
	print(str1)
	books_df.loc[i,'b_slug'] = str1

print(books_df.columns)
'''

'''
headings = list(publishers_df.columns)
print(headings[0])

comands = ''
for i in range(publishers_df.shape[0]):
	one_command_str = 'Publisher.objects.create('
	for j in range(publishers_df.shape[1]):
		if publishers_df.iloc[i][headings[j]] != None :
			one_command_str = one_command_str + headings[j] + '=\"' + str(publishers_df.iloc[i][headings[j]]) + '\",'
	one_command_str += ')\n'
	comands += one_command_str
print(comands)
'''
